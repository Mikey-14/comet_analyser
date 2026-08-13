"""
主窗口模块
========
彗星实验分析工具的主界面 - 交互式半自动分析工作流。

工作流：
1. 加载图片 → 2. 框选背景区域 → 3. 逐一框选细胞 → 4. 切换图片重复 → 5. 导出Excel
"""

import os
import cv2
import numpy as np
from datetime import datetime
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QAction, QToolBar, QMenu,
    QFileDialog, QMessageBox, QStatusBar, QSplitter,
    QListWidget, QGroupBox, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView,
    QFrame, QSizePolicy, QInputDialog
)
from PyQt5.QtCore import Qt, QSize, QRect
from PyQt5.QtGui import QPixmap, QImage

from analysis.preprocess import load_image, denoise, enhance_clahe
from analysis.segmentation import segment_comet
from analysis.metrics import compute_all_metrics
from .image_canvas import ImageCanvas


# ==================== 颜色常量 ====================
COLOR_BG = (255, 50, 50)
COLOR_CELL = (50, 200, 50)


class MainWindow(QMainWindow):
    """彗星实验分析主窗口 - 交互式工作流"""

    MODE_IDLE = "idle"
    MODE_BG = "background"
    MODE_CELL = "cell"

    def __init__(self):
        super().__init__()
        self.setWindowTitle("彗星实验分析工具 - Comet Assay Analyser")
        self.setMinimumSize(1200, 800)

        # ---- 核心状态 ----
        self.mode = self.MODE_IDLE
        self.current_image_path: str = None
        self._raw_cv_image: np.ndarray = None
        self.bg_mean: float = None
        self.current_cells: list = []
        self.all_results: list = []
        self.image_files: list = []
        self._cell_counter = 0
        self.pixel_size_um: float = 1.0

        self._init_ui()
        self._create_menu_bar()
        self._create_toolbar()
        self._create_status_bar()
        self._update_mode_ui()

    # ==================== UI 初始化 ====================

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(4, 4, 4, 4)

        # ====== 左侧面板 ======
        left_panel = QGroupBox("图片列表")
        left_layout = QVBoxLayout(left_panel)

        self.file_list = QListWidget()
        self.file_list.itemClicked.connect(self._on_file_selected)
        self.file_list.setMaximumWidth(200)

        btn_load = QPushButton("加载图片...")
        btn_load.clicked.connect(self._on_load_image)

        nav_layout = QHBoxLayout()
        self.btn_prev = QPushButton("< 上一张")
        self.btn_prev.clicked.connect(self._on_prev_image)
        self.btn_next = QPushButton("下一张 >")
        self.btn_next.clicked.connect(self._on_next_image)
        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.btn_next)

        self.lbl_image_index = QLabel("")
        self.lbl_image_index.setAlignment(Qt.AlignCenter)

        left_layout.addWidget(btn_load)
        left_layout.addWidget(self.file_list)
        left_layout.addLayout(nav_layout)
        left_layout.addWidget(self.lbl_image_index)

        # ====== 中间面板 ======
        image_panel = QWidget()
        image_layout = QVBoxLayout(image_panel)

        mode_layout = QHBoxLayout()
        self.btn_bg = QPushButton("框选背景区域")
        self.btn_bg.setCheckable(True)
        self.btn_bg.clicked.connect(lambda: self._enter_mode(self.MODE_BG))
        self.btn_bg.setToolTip("在图片上拖拽框选一片没有细胞的背景区域")
        self.btn_bg.setStyleSheet(
            "QPushButton:checked { background-color: #f8d7da; border: 2px solid #dc3545; }"
        )

        self.btn_cell = QPushButton("框选细胞")
        self.btn_cell.setCheckable(True)
        self.btn_cell.clicked.connect(lambda: self._enter_mode(self.MODE_CELL))
        self.btn_cell.setToolTip("在图片上拖拽框选一个彗星细胞区域")
        self.btn_cell.setStyleSheet(
            "QPushButton:checked { background-color: #d4edda; border: 2px solid #28a745; }"
        )

        self.btn_done = QPushButton("完成当前图片")
        self.btn_done.clicked.connect(self._on_finish_image)
        self.btn_done.setToolTip("保存当前图片结果，切换到下一张")
        self.btn_done.setStyleSheet("QPushButton { background-color: #cce5ff; }")

        mode_layout.addWidget(self.btn_bg)
        mode_layout.addWidget(self.btn_cell)
        mode_layout.addStretch()
        mode_layout.addWidget(self.btn_done)

        self.canvas = ImageCanvas()
        self.canvas.setMinimumSize(500, 400)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.canvas.setAlignment(Qt.AlignCenter)
        self.canvas.setStyleSheet(
            "ImageCanvas { background-color: #e8e8e8; border: 1px solid #ccc; }"
        )
        self.canvas.rect_selected.connect(self._on_rect_selected)

        self.lbl_mode_hint = QLabel("请先加载图片，然后点击「框选背景区域」")
        self.lbl_mode_hint.setStyleSheet(
            "QLabel { font-size: 13px; padding: 6px; background: #fff3cd; "
            "border-radius: 4px; color: #856404; }"
        )

        image_layout.addLayout(mode_layout)
        image_layout.addWidget(self.canvas, stretch=1)
        image_layout.addWidget(self.lbl_mode_hint)

        # ====== 右侧面板 ======
        right_panel = QGroupBox("分析结果")
        right_layout = QVBoxLayout(right_panel)
        right_panel.setMinimumWidth(300)
        right_panel.setMaximumWidth(380)

        lbl_current = QLabel("<b>当前图片细胞：</b>")
        self.cell_table = QTableWidget()
        self.cell_table.setColumnCount(2)
        self.cell_table.setHorizontalHeaderLabels(["细胞", "Tail DNA%"])
        self.cell_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.cell_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.cell_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.cell_table.setAlternatingRowColors(True)
        self.cell_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.cell_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.cell_table.customContextMenuRequested.connect(self._on_table_context_menu)

        self.btn_delete_cell = QPushButton("删除选中细胞")
        self.btn_delete_cell.clicked.connect(self._on_delete_cell)
        self.btn_delete_cell.setEnabled(False)

        right_layout.addWidget(lbl_current)
        right_layout.addWidget(self.cell_table)
        right_layout.addWidget(self.btn_delete_cell)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        right_layout.addWidget(line)

        self.lbl_summary = QLabel("<b>全局汇总：</b>尚未分析任何细胞")
        self.lbl_summary.setWordWrap(True)
        right_layout.addWidget(self.lbl_summary)

        self.btn_export = QPushButton("导出 Excel")
        self.btn_export.clicked.connect(self._on_export_excel)
        self.btn_export.setEnabled(False)
        self.btn_export.setStyleSheet(
            "QPushButton { background-color: #28a745; color: white; font-weight: bold; }"
        )
        right_layout.addWidget(self.btn_export)

        # ====== 分割器 ======
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(image_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setStretchFactor(2, 0)

        main_layout.addWidget(splitter)

    # ==================== 菜单栏 ====================

    def _create_menu_bar(self):
        menu = self.menuBar()
        file_menu = menu.addMenu("文件(&F)")

        open_action = QAction("打开图片(&O)...", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._on_load_image)
        file_menu.addAction(open_action)

        open_dir_action = QAction("打开文件夹(&D)...", self)
        open_dir_action.setShortcut("Ctrl+Shift+O")
        open_dir_action.triggered.connect(self._on_load_folder)
        file_menu.addAction(open_dir_action)

        file_menu.addSeparator()

        export_action = QAction("导出 Excel...", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self._on_export_excel)
        file_menu.addAction(export_action)

        file_menu.addSeparator()
        exit_action = QAction("退出(&Q)", self)
        exit_action.setShortcut("Alt+F4")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        settings_menu = menu.addMenu("设置(&S)")
        pixel_action = QAction("像素尺寸...", self)
        pixel_action.triggered.connect(self._on_set_pixel_size)
        settings_menu.addAction(pixel_action)

        help_menu = menu.addMenu("帮助(&H)")
        about_action = QAction("关于(&A)", self)
        about_action.triggered.connect(self._on_about)
        help_menu.addAction(about_action)

    # ==================== 工具栏 ====================

    def _create_toolbar(self):
        toolbar = QToolBar("主工具栏")
        toolbar.setIconSize(QSize(24, 24))
        toolbar.setMovable(False)
        self.addToolBar(toolbar)
        toolbar.addWidget(QPushButton("打开图片", clicked=self._on_load_image))
        toolbar.addSeparator()
        toolbar.addWidget(QPushButton("<", clicked=self._on_prev_image, maximumWidth=40))
        toolbar.addWidget(QPushButton(">", clicked=self._on_next_image, maximumWidth=40))
        toolbar.addSeparator()
        toolbar.addWidget(QPushButton("导出 Excel", clicked=self._on_export_excel))

    # ==================== 状态栏 ====================

    def _create_status_bar(self):
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪 — 请加载彗星实验图片")

    # ==================== 模式切换 ====================

    def _enter_mode(self, new_mode: str):
        if self._raw_cv_image is None:
            QMessageBox.information(self, "提示", "请先加载一张图片")
            self.btn_bg.setChecked(False)
            self.btn_cell.setChecked(False)
            return

        if self.mode == new_mode:
            self.mode = self.MODE_IDLE
            self.btn_bg.setChecked(False)
            self.btn_cell.setChecked(False)
            self.canvas.setCursor(Qt.ArrowCursor)
            self._update_mode_ui()
            return

        self.mode = new_mode
        self.btn_bg.setChecked(new_mode == self.MODE_BG)
        self.btn_cell.setChecked(new_mode == self.MODE_CELL)
        self.canvas.clear_selection()

        if new_mode == self.MODE_BG:
            self.canvas.set_selection_color(*COLOR_BG)
            self.canvas.setCursor(Qt.CrossCursor)
            self.lbl_mode_hint.setText("背景模式：请在图像上拖拽框选一片 <b>没有细胞的背景区域</b>")
            self.lbl_mode_hint.setStyleSheet(
                "QLabel { font-size: 13px; padding: 6px; background: #fce4e4; "
                "border-radius: 4px; color: #721c24; }"
            )
        elif new_mode == self.MODE_CELL:
            if self.bg_mean is None:
                QMessageBox.warning(self, "提示", "请先框选背景区域，再框选细胞！")
                self.mode = self.MODE_IDLE
                self.btn_cell.setChecked(False)
                self._update_mode_ui()
                return
            self.canvas.set_selection_color(*COLOR_CELL)
            self.canvas.setCursor(Qt.CrossCursor)
            self.lbl_mode_hint.setText("细胞模式：请在图像上拖拽框选 <b>一个彗星细胞</b>（右键取消）")
            self.lbl_mode_hint.setStyleSheet(
                "QLabel { font-size: 13px; padding: 6px; background: #d4edda; "
                "border-radius: 4px; color: #155724; }"
            )
        else:
            self._update_mode_ui()

        self.status_bar.showMessage(f"当前模式: {self.mode}")

    def _update_mode_ui(self):
        if self.mode == self.MODE_IDLE:
            if self._raw_cv_image is None:
                self.lbl_mode_hint.setText("请先加载图片")
                self.lbl_mode_hint.setStyleSheet(
                    "QLabel { font-size: 13px; padding: 6px; background: #fff3cd; "
                    "border-radius: 4px; color: #856404; }"
                )
            elif self.bg_mean is None:
                self.lbl_mode_hint.setText("请点击「框选背景区域」开始")
                self.lbl_mode_hint.setStyleSheet(
                    "QLabel { font-size: 13px; padding: 6px; background: #fff3cd; "
                    "border-radius: 4px; color: #856404; }"
                )
            else:
                self.lbl_mode_hint.setText("请点击「框选细胞」继续添加细胞")
                self.lbl_mode_hint.setStyleSheet(
                    "QLabel { font-size: 13px; padding: 6px; background: #d4edda; "
                    "border-radius: 4px; color: #155724; }"
                )

    # ==================== 图片加载 ====================

    def _on_load_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择彗星实验图片", "",
            "图片文件 (*.png *.jpg *.jpeg *.bmp *.tif *.tiff);;所有文件 (*)"
        )
        if path:
            self._save_current_results()
            self.all_results = []
            self.image_files = [path]
            self.file_list.clear()
            self.file_list.addItem(os.path.basename(path))
            self._switch_to_image(path)

    def _on_load_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择包含彗星图片的文件夹")
        if not folder:
            return

        self._save_current_results()
        self.all_results = []

        supported = ('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')
        self.image_files = []
        self.file_list.clear()
        for f in sorted(os.listdir(folder)):
            if f.lower().endswith(supported):
                full = os.path.join(folder, f)
                self.image_files.append(full)
                self.file_list.addItem(f)

        self.status_bar.showMessage(f"已加载 {len(self.image_files)} 张图片")
        if self.image_files:
            self._switch_to_image(self.image_files[0])

    def _on_file_selected(self, item):
        idx = self.file_list.row(item)
        if 0 <= idx < len(self.image_files):
            self._save_current_results()
            self._switch_to_image(self.image_files[idx])

    def _on_prev_image(self):
        if not self.image_files:
            return
        self._save_current_results()
        current_idx = self._get_current_index()
        new_idx = (current_idx - 1) % len(self.image_files)
        self._switch_to_image(self.image_files[new_idx])

    def _on_next_image(self):
        if not self.image_files:
            return
        self._save_current_results()
        current_idx = self._get_current_index()
        new_idx = (current_idx + 1) % len(self.image_files)
        self._switch_to_image(self.image_files[new_idx])

    def _get_current_index(self) -> int:
        if not self.current_image_path or not self.image_files:
            return 0
        try:
            return self.image_files.index(self.current_image_path)
        except ValueError:
            return 0

    def _switch_to_image(self, path: str):
        self.current_image_path = path
        try:
            self._raw_cv_image = load_image(path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法加载图片:\n{e}")
            return

        height, width = self._raw_cv_image.shape[:2]
        if len(self._raw_cv_image.shape) == 2:
            qimg = QImage(self._raw_cv_image.data, width, height, width,
                          QImage.Format_Grayscale8)
        else:
            rgb = cv2.cvtColor(self._raw_cv_image, cv2.COLOR_BGR2RGB)
            qimg = QImage(rgb.data, width, height, width * 3,
                          QImage.Format_RGB888)

        pixmap = QPixmap.fromImage(qimg)
        self.canvas.set_image(pixmap)

        # 重置图片级状态
        self.bg_mean = None
        self.current_cells = []
        self._cell_counter = 0
        self.mode = self.MODE_IDLE
        self.btn_bg.setChecked(False)
        self.btn_cell.setChecked(False)
        self._update_mode_ui()
        self._update_cell_table()

        if self.image_files:
            idx = self._get_current_index()
            self.lbl_image_index.setText(f"{idx + 1} / {len(self.image_files)}")
            self.file_list.setCurrentRow(idx)

        self.status_bar.showMessage(f"当前图片: {os.path.basename(path)}")

        # 恢复之前暂存的结果
        for result in self.all_results:
            if result.get("path") == path:
                self._restore_results(result)
                break

    def _save_current_results(self):
        if not self.current_image_path:
            return
        path = self.current_image_path

        if not self.current_cells:
            self.all_results = [r for r in self.all_results if r["path"] != path]
            self._update_global_summary()
            return

        for i, result in enumerate(self.all_results):
            if result["path"] == path:
                self.all_results[i] = {"path": path, "filename": os.path.basename(path), "cells": self.current_cells.copy()}
                self._update_global_summary()
                return
        self.all_results.append({"path": path, "filename": os.path.basename(path), "cells": self.current_cells.copy()})
        self._update_global_summary()

    def _restore_results(self, result: dict):
        self.current_cells = result.get("cells", [])
        self._cell_counter = len(self.current_cells)
        if self.current_cells:
            self.bg_mean = self.current_cells[0].get("bg_mean", None)
        self._update_cell_table()
        self._update_mode_ui()

    # ==================== 框选回调（核心） ====================

    def _on_rect_selected(self, rect: QRect):
        if self._raw_cv_image is None:
            return

        roi = self.canvas.get_selected_region_numpy(self._raw_cv_image)
        if roi is None or roi.size == 0:
            return

        if self.mode == self.MODE_BG:
            self._handle_background_selection(roi)
        elif self.mode == self.MODE_CELL:
            self._handle_cell_selection(roi)

    def _handle_background_selection(self, roi: np.ndarray):
        if len(roi.shape) == 3:
            gray_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        else:
            gray_roi = roi
        self.bg_mean = float(np.mean(gray_roi))

        self.mode = self.MODE_IDLE
        self.btn_bg.setChecked(False)
        self.canvas.clear_selection()
        self._update_mode_ui()

        if self.current_cells:
            QMessageBox.warning(
                self, "背景已更新",
                f"背景区域平均强度已更新为: <b>{self.bg_mean:.1f}</b>\n\n"
                f"⚠ 注意：已有 {len(self.current_cells)} 个细胞的分析结果不会重新计算，"
                f"仍保留旧背景值。新添加的细胞将使用新背景。"
            )
        else:
            QMessageBox.information(
                self, "背景已设置",
                f"背景区域平均强度: <b>{self.bg_mean:.1f}</b>\n\n"
                f"现在可以开始框选细胞了！"
            )
        self.status_bar.showMessage(f"背景已设置 — 平均强度: {self.bg_mean:.1f}")

    def _handle_cell_selection(self, roi: np.ndarray):
        try:
            if len(roi.shape) == 3:
                raw_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            else:
                raw_gray = roi.copy()

            # 背景减除 + 预处理（用于分割）
            if self.bg_mean is not None:
                processed = cv2.subtract(raw_gray, self.bg_mean)
                processed = cv2.normalize(processed, None, 0, 255, cv2.NORM_MINMAX)
                processed = processed.astype(np.uint8)
            else:
                processed = raw_gray.copy()

            # 去噪 + CLAHE
            processed = denoise(processed, method="gaussian", kernel_size=3)
            processed = enhance_clahe(processed, clip_limit=1.5)

            # 分割
            region = segment_comet(processed)
            if region is None:
                QMessageBox.warning(
                    self, "分析失败",
                    "未能在此区域中检测到彗星。\n请检查框选是否包含完整的彗星细胞。"
                )
                return

            # 指标计算：使用背景扣除后的原图（不归一化，保留真实荧光强度比例）
            if self.bg_mean is not None:
                gray_bg_corrected = cv2.subtract(raw_gray, self.bg_mean)
            else:
                gray_bg_corrected = raw_gray
            metrics = compute_all_metrics(gray_bg_corrected, region,
                                          pixel_size_um=self.pixel_size_um)

            # 记录细胞
            self._cell_counter += 1
            label = f"Cell_{self._cell_counter}"
            self.current_cells.append({
                "label": label,
                "metrics": metrics,
                "bg_mean": self.bg_mean
            })

            # 更新UI
            self._save_current_results()
            self._update_cell_table()

            self.status_bar.showMessage(
                f"{label} 分析完成 — Tail DNA%: {metrics.tail_dna_percent:.1f}%, "
                f"Tail Length: {metrics.tail_length_px:.1f} px"
            )

            # 清除框选以便继续，但保持细胞模式
            self.canvas.clear_selection()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"分析过程出错:\n{str(e)}")

    # ==================== 完成当前图片 ====================

    def _on_finish_image(self):
        if not self.current_cells:
            QMessageBox.information(self, "提示", "当前图片还没有分析任何细胞")
            return

        self._save_current_results()

        idx = self._get_current_index()
        if idx < len(self.image_files) - 1:
            self._switch_to_image(self.image_files[idx + 1])
        else:
            QMessageBox.information(
                self, "全部完成",
                "所有图片已分析完毕！\n请点击「导出 Excel」保存结果。"
            )

    # ==================== 结果表格 ====================

    def _update_cell_table(self):
        self.cell_table.setRowCount(len(self.current_cells))
        for i, cell in enumerate(self.current_cells):
            m = cell["metrics"]
            self.cell_table.setItem(i, 0, QTableWidgetItem(cell["label"]))
            dna_item = QTableWidgetItem(f"{m.tail_dna_percent:.1f}%")
            dna_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.cell_table.setItem(i, 1, dna_item)
        self.btn_delete_cell.setEnabled(len(self.current_cells) > 0)

    def _on_delete_cell(self):
        row = self.cell_table.currentRow()
        if 0 <= row < len(self.current_cells):
            del self.current_cells[row]
            self._cell_counter = len(self.current_cells)
            for i, cell in enumerate(self.current_cells):
                cell["label"] = f"Cell_{i + 1}"
            self._save_current_results()
            self._update_cell_table()

    def _on_table_context_menu(self, pos):
        row = self.cell_table.rowAt(pos.y())
        if 0 <= row < len(self.current_cells):
            menu = QMenu()
            action = menu.addAction("删除此细胞")
            if menu.exec_(self.cell_table.mapToGlobal(pos)) == action:
                self.cell_table.selectRow(row)
                self._on_delete_cell()

    def _update_global_summary(self):
        total_cells = sum(len(r["cells"]) for r in self.all_results)
        img_count = len(self.all_results)
        if total_cells == 0:
            self.lbl_summary.setText("<b>全局汇总：</b>尚未分析任何细胞")
            self.btn_export.setEnabled(False)
        else:
            self.lbl_summary.setText(
                f"<b>全局汇总：</b><br>"
                f"已分析图片: {img_count} 张<br>"
                f"已分析细胞: {total_cells} 个"
            )
            self.btn_export.setEnabled(True)

    # ==================== 导出 Excel ====================

    def _on_export_excel(self):
        self._save_current_results()

        if not self.all_results:
            QMessageBox.information(self, "提示", "没有可导出的结果")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "导出 Excel",
            f"comet_results_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "细胞分析结果"

            headers = [
                "图片文件名", "细胞编号",
                "Tail Length (px)", "Tail Length (um)",
                "Head Area (px)", "Tail Area (px)", "Comet Area (px)",
                "Head Intensity", "Tail Intensity", "Comet Intensity",
                "Tail DNA%", "Tail Moment", "Olive Tail Moment"
            ]

            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')

            row = 2
            for result in self.all_results:
                fname = result["filename"]
                for cell_data in result["cells"]:
                    m = cell_data["metrics"]
                    vals = [
                        fname, cell_data["label"],
                        round(m.tail_length_px, 2), round(m.tail_length_um, 2),
                        m.head_area_px, m.tail_area_px, m.comet_area_px,
                        round(m.head_intensity, 2), round(m.tail_intensity, 2),
                        round(m.comet_intensity, 2),
                        round(m.tail_dna_percent, 2),
                        round(m.tail_moment, 2),
                        round(m.olive_tail_moment, 2)
                    ]
                    for col, v in enumerate(vals, 1):
                        ws.cell(row=row, column=col, value=v)
                    row += 1

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

            # Sheet 2: 汇总
            ws2 = wb.create_sheet("图片汇总")
            sum_headers = ["图片文件名", "细胞数量", "平均 Tail DNA%", "平均 Tail Length (px)", "平均 Olive Tail Moment"]
            for col, h in enumerate(sum_headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = hf
                cell.fill = hfill
                cell.alignment = Alignment(horizontal='center')

            row2 = 2
            for result in self.all_results:
                cells = result["cells"]
                if not cells:
                    continue
                n = len(cells)
                avg_dna = np.mean([c["metrics"].tail_dna_percent for c in cells])
                avg_len = np.mean([c["metrics"].tail_length_px for c in cells])
                avg_otm = np.mean([c["metrics"].olive_tail_moment for c in cells])
                for col, v in enumerate([
                    result["filename"], n,
                    round(avg_dna, 2), round(avg_len, 2), round(avg_otm, 2)
                ], 1):
                    ws2.cell(row=row2, column=col, value=v)
                row2 += 1

            for col in range(1, len(sum_headers) + 1):
                ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 22

            wb.save(path)
            self.status_bar.showMessage(f"已导出: {path}")
            QMessageBox.information(self, "导出成功", f"结果已保存到:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    # ==================== 像素尺寸设置 ====================

    def _on_set_pixel_size(self):
        current = self.pixel_size_um
        value, ok = QInputDialog.getDouble(
            self, "像素尺寸设置",
            "请输入每个像素对应的微米数：\n"
            "（通过显微镜标定尺测量得出）",
            current, 0.01, 100.0, 3
        )
        if ok:
            self.pixel_size_um = value
            self.status_bar.showMessage(f"像素尺寸已设置为: {value:.3f} µm/px")

    # ==================== 窗口关闭 ====================

    def closeEvent(self, event):
        self._save_current_results()
        if self.all_results:
            reply = QMessageBox.question(
                self, "确认退出",
                "当前有未导出的分析结果。\n是否在退出前导出为 Excel？",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            if reply == QMessageBox.Yes:
                self._on_export_excel()
                event.accept()
            elif reply == QMessageBox.No:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    # ==================== 关于 ====================

    def _on_about(self):
        QMessageBox.about(
            self,
            "关于 Comet Analyser",
            "<h3>彗星实验分析工具 v2.0</h3>"
            "<p>交互式半自动彗星实验 (Comet Assay) 图像分析工具。</p>"
            "<p><b>工作流：</b></p>"
            "<ol>"
            "<li>加载图片</li>"
            "<li>框选背景区域</li>"
            "<li>逐一框选细胞进行分析</li>"
            "<li>切换图片继续</li>"
            "<li>导出 Excel 结果</li>"
            "</ol>"
            "<p><b>指标：</b>Tail Length · Tail DNA% · Tail Moment · Olive Tail Moment</p>"
            "<p>基于 PyQt5 + OpenCV + scikit-image 构建</p>"
        )
